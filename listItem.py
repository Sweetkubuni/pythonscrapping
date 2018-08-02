from pywinauto.application import Application
from marionette_driver.marionette import Marionette
from marionette_driver import By
from multiprocessing import Process, Queue
from openpyxl import load_workbook
import subprocess
import time
import re
import os
import errno
import urllib
import sys

def startbroswer():
    subprocess.call(['firefox', '-marionette'])

def getPID():
    found = subprocess.check_output('tasklist /v /fo csv | findstr /i "firefox"', shell=True)
    process_list = [x for x in (y.split() for y in found.splitlines()) if x]
    return int(process_list[0][0].split(',')[1].replace('\"',""))

def parse_index( text ):
    print "parsing for index on text:" + text + "\n"
    m = re.search('[0-9]+', text)
    return int(m.group(0))

def check_empty( s ):
	if s == "":
		return -1
	else:
		return s

def parse_quality( text ):
    mtext = text.upper()
    if mtext == "NEW":
        return 1
    elif mtext == "LIKE NEW":
        return 2
    elif mtext == "GOOD":
        return 3
    elif mtext == "FAIR":
        return 4
    elif mtext == "POOR":
        return 5

def download_images(ws, len, row):
    cell = lambda c, r: '{0}{1}'.format(c, r + 1)
    len = int(ws[cell('L', row)].value)
    try:
        os.makedirs('photos')
    except OSError as e:
        if e.errno != errno.EEXIST:
            raise
    for i in range(len):
        letter = chr(ord('N') + i)
        print "getting photo from the internet and saving at " + "photos/img" + str(i) + ".jpg\n"
        print urllib.urlretrieve(ws[cell(letter, row)].value, "photos/img" + str(i) + ".jpg")

def load_excel(filename, rowlen):
    wb = load_workbook(filename)
    ws = wb.active
    data = []
    cell = lambda c, r: '{0}{1}'.format(c, r + 1)
    for i in range(0, rowlen):
        x = i + 1
        data.append((ws[cell('B', x)].value, ws[cell('C', x)].value, parse_index(ws[cell('D', x)].value), parse_index(ws[cell('E', x)].value), parse_index(ws[cell('F', x)].value), int(check_empty(ws[cell('G', x)].value)), ws[cell('H', x)].value, parse_quality(ws[cell('I', x)].value), int(ws[cell('J', x)].value), ws[cell('K', x)].value, int(ws[cell('L', x)].value), int(ws[cell('M', x)].value)))
        print data[i]
        print "\n\n"
        download_images(ws, int(ws[cell('L', x)].value), x)
    return data
def fileUpload(client, pid, path, index = 0):
    if index > 0:
        client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[1]/div/ul/li[{}]/div/div".format(index + 1)).click()
    else:
        client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[1]/div/ul/li/div/div").click()
    time.sleep(1)
    app = Application().connect(process=pid)
    dlg = app.top_window()
    print "submitting path into fileopen:" + path + "\n"
    dlg["Edit"].set_text(path)
    time.sleep(2)
    dlg["Open"].click()
    try:
        app.Dialog.Open.Click()
    except:
        return
    time.sleep(2)

def select_option(row, choice):
	return "/html/body/div[1]/main/div[3]/div/div[2]/div[4]/div/div[2]/div/div[1]/div/div/div[2]/div[{}]/select/option[{}]".format(row, choice)


def category(client, sel1, sel2, sel3):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[4]/div/div").click()
    time.sleep(2)
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[4]/div/div[2]/div/div[1]/div/div/div/div").click()
    select = []
    select.extend((sel1,sel2, sel3))
    for i,sel in enumerate(select):
        print "i:" + str(i) + " sel:" + str(sel) + "\n"
        client.find_element(By.XPATH, select_option(i + 1, sel + 1)).click()
        time.sleep(2)

def determineSize(client, size):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[4]/div[2]/div/div[2]/div/select/option[{}]".format(size)).click()

def select_new(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[5]/div/div[2]/label[1]/div").click()
def select_likeNew(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[5]/div/div[2]/label[2]/div").click()
def select_good(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[5]/div/div[2]/label[3]/div").click()
def select_fair(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[5]/div/div[2]/label[4]/div").click()
def select_poor(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[5]/div/div[2]/label[5]/div").click()

def selectCondition(client, size):
    switch = {
        1 : select_new,
        2 : select_likeNew,
        3 : select_good,
        4 : select_fair,
        5 : select_poor,
    }
    switch[size](client)


def ship_selfpaid(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/ul/li[2]/div/label/input").click()

def select_weight(weight):
	return "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[3]/div/div/div/div[{}]/label/div/div".format(weight)

def ship_prepaid(client, who_pays, weight, shipping_method):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/ul/li[1]/div/label/input").click()
    if who_pays is "seller":
        client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div/div/div[1]/label/input").click()
    elif who_pays is "customer":
        client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div/div/div[2]/label/input").click()
    select_weight(weight)
    if weight < 5:
        if shipping_method is "USPS":
            client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[4]/div/div/div[1]/label/input").click()
        elif shipping_method is "FEDEX":
            client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[4]/div/div/div[2]/label/input").click()
        else:
            print("Error: Select wrong shipping method\n")
    else:
        client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/div[4]/div/div/div/label/input").click()
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[2]/div[2]/div[1]/div/div[2]/button").click()

def submit_button(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[8]/div").click()
def submit_new_item(client):
    client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[3]/div[1]/div/div/div/a").click()
def submit_size(client, size):

def workerProc(arg):
    client = Marionette('localhost', port=2828)
    client.start_session()
    client.navigate('https://www.mercari.com/sell/')
    time.sleep(5)
    #first open the excel file to grab information
    #photos will be stored in photos file
    num_rows = arg
    items = load_excel("Mercari.xlsx", num_rows)

    for item in items:
        title = client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[3]/div[1]/div[2]/input")
        title.send_keys(item[0])
        textarea = client.find_element(By.TAG_NAME, 'textarea')
        textarea.send_keys(item[1])
        zipcode = client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[6]/div[1]/div[2]/input")
        zipcode.send_keys(item[8])
        price = client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[7]/div/div[2]/input")
        price.send_keys(item[10])
        brand = client.find_element(By.XPATH, "/html/body/div[1]/main/div[3]/div/div[2]/div[7]/div/div[2]/input")
        brand.send_keys(item[6])
        category(client, item[2], item[3], item[4])

        if item[6] >= 0:
            submit_size(client, item[6])

        time.sleep(1)
        #determineSize(client, 1)
        selectCondition(client, item[7])
        ship_selfpaid(client)


        for i in range(item[10]):
            path = os.path.dirname(os.path.realpath("photos\img" + str(i) + ".jpg"))
            fileUpload(client, getPID(), path + "\img" + str(i) + ".jpg", i)
        submit_button(client)
        time.sleep(2)
        client.navigate('https://www.mercari.com/sell/')


if __name__ == "__main__":
    #if len(sys.argv) != 2 :
        #print "missing argument for number of items"
        #sys.exit(1)
    #num_rows = sys.argv[1]
    browser   = Process(name="browser", target=startbroswer)
    worker    = Process(name="workerProc", target=workerProc, args=(1,))
    browser.start()
    worker.start()
    worker.join()
    browser.join()
